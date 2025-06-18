import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
print("✅ openpyxl imported successfully")

st.set_page_config(page_title="ERP Column Mapper", layout="centered")
st.title("ERP Column Mapper and Transfer Tool")

# Step 1: Upload source file
st.header("Step 1: Upload Source (Consumption) Excel File")
source_file = st.file_uploader("Upload Source File", type=["xlsx"], key="source")

# Step 2: Upload target file
st.header("Step 2: Upload Target (Transfer) Excel File")
target_file = st.file_uploader("Upload Transfer File", type=["xlsx"], key="target")

# Define column mapping
column_mapping = {
    'UTC Date & Time': 'TIMESTAMP',
    'Event': 'EVENT_TYPE',
    'From Port': 'PORT',
    'Steaming time (HRS)': 'DUR_SEA',
    'Obs distance (NM)': 'DIST',
    'Time Spent at Anchorage (Hrs)': 'DUR_ANC',
    'Time Spent at Drifting (Hrs)': 'DUR_DRIFT',
    'Total cargo on board (MT)': 'CARGO_MT',
    'AE LS MGO consumption (MT)': 'AE_MGO_CONS',
    'ME LS MGO consumption (MT)': 'ME_MGO_CONS',
    'BLR LS MGO consumption (MT)': 'Boiler_MGO_CONS',
    'AE VLSFO consumption (MT)': 'AE_HFO_CONS',
    'ME VLSFO consumption (MT)': 'ME_HFO_CONS',
    'BLR VLSFO consumption (MT)': 'Boiler_HFO_CONS',
    'ROB LS MGO': 'MGO_ROB',
    'ROB VLSFO': 'HFO_ROB',
}

if source_file and target_file:
    source_df = pd.read_excel(source_file)

    # Check for missing columns
    missing_cols = [col for col in column_mapping if col not in source_df.columns]
    if missing_cols:
        st.error(f"Missing columns in source file: {missing_cols}")
    else:
        # Map columns and fill NaNs with 0
        mapped_df = source_df[list(column_mapping.keys())].rename(columns=column_mapping)
        mapped_df = mapped_df.fillna(0)

        st.subheader("Mapped Data Preview")
        st.dataframe(mapped_df)

        # Load target workbook and sheet
        wb = openpyxl.load_workbook(target_file)
        if 'VOYAGE' not in wb.sheetnames:
            st.error("The sheet named 'VOYAGE' does not exist in the target file.")
        else:
            ws = wb['VOYAGE']

            # Get the headers in the VOYAGE sheet
            header_row = [cell.value for cell in ws[1]]
            col_index_map = {col_name: idx + 1 for idx, col_name in enumerate(header_row)}

            # Write mapped data into matching columns only
            for row_num, row_data in enumerate(mapped_df.itertuples(index=False), start=2):
                for col_name, value in zip(mapped_df.columns, row_data):
                    if col_name in col_index_map:
                        col_index = col_index_map[col_name]
                        if pd.isna(value):  # Fallback safety check
                            value = 0
                        ws.cell(row=row_num, column=col_index, value=value)

            # Save to in-memory file
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success("Mapped data successfully written into the 'VOYAGE' sheet with missing values filled as 0.")

            st.download_button(
                label="Download Updated File",
                data=output.getvalue(),
                file_name="Mapped_VOYAGE_File.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
